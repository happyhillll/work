def ws_to_dict(ws):
    datas = [[cell.value for cell in row] for row in ws.rows]  # [[1, 'A'], [2, 'B'], [3, 'C'], [4, 'D'], [5, 'E']]
    return {
        d[0]: d[1] for d in datas  # {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E'}
    }


def use_openpyxl(paths):
    from openpyxl import load_workbook
    ws_1, ws_2 = load_workbook(paths[0])['Sheet1'], load_workbook(paths[1])['Sheet1']
    dict_1, dict_2 = ws_to_dict(ws_1), ws_to_dict(ws_2)
    output = {
        k: dict_2[v] for k, v in dict_1.items()  # {1: 'ㄱ', 2: 'ㄴ', 3: 'ㄷ', 4: 'ㄹ', 5: 'ㅁ'}
    }

    # TODO write output.xlsx


if __name__ == "__main__":
    paths = [
        r"C:\Users\cheolhunheo\Desktop\projects\temp\data\input1.xlsx",
        r"C:\Users\cheolhunheo\Desktop\projects\temp\data\input2.xlsx",
        r"C:\Users\cheolhunheo\Desktop\projects\temp\data\output.xlsx"
    ]
    use_openpyxl(paths)
