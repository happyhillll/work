import openpyxl

# 1. 엑셀 파일 열기
workbook = openpyxl.load_workbook('data.xlsx')
# 현재 활성 중인 워크시트를 선택하는 방법
worksheet = workbook.active

# 2. 데이터 읽어서 딕셔너리 만들기
result_dic = {}
for row in worksheet.iter_rows(min_row=1, values_only=True):
    key = str(row[0]) + "#" + str(row[1])
    values = str(row[2]).split('\n') if row[2] else ['']
    result_dic[key] = values

# 3. 딕셔너리 출력하기
print(result_dic)

# 4. 결과 엑셀 파일 쓰기
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active
for key, values in result_dic.items():
    for value in values:
        if value:
            output_worksheet.append([key, value])
        else:
            output_worksheet.append([key, ""])

output_workbook.save('output.xlsx')